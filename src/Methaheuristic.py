import random
import math
from openpyxl import workbook
from pathlib import Path
from numpy import cos, pi, exp, sqrt, prod, sin
import numpy as np
import string
import openpyxl as op
import argparse

class Functions:

    def __init__(self,type):
        self.type=type


    def calculaFitness(self,number):
        if (self.type == "1"):
            y=[number for x in range(30)]
            return self.F1(y)
        elif (self.type == "2"):
            y = [number for x in range(30)]
            return self.F2(y)
        elif (self.type == "3"):
            y = [number for x in range(30)]
            return self.F3(y)
        elif (self.type == "4"):
            y = [number for x in range(30)]
            return self.F4(y)
        elif (self.type == "5"):
            y = [number for x in range(30)]
            return self.F5(y)
        elif (self.type == "6"):
            y = [number for x in range(30)]
            return self.F6(y)
        elif (self.type == "7"):
            y = [number for x in range(30)]
            return self.F7(y)
        elif (self.type == "8"):
            y = [number for x in range(30)]
            return self.F8(y)
        elif (self.type == "9"):
            y = [number for x in range(30)]
            return self.F9(y)
        elif (self.type == "10"):
            y = [number for x in range(2)]
            return self.F10(y)
        elif (self.type == "11"):
            y = [number for x in range(2)]
            return self.F11(y)
        elif (self.type == "12"):
            return self.F12(number)
        elif (self.type == "13"):
            y = [number for x in range(2)]
            return self.F13(y)
        elif self.type == "14":
            y = [number for x in range(3)]
            return self.F14(y)
        else:
            y = [number for x in range(6)]
            return self.F15(y)

    def F1(self,x):
        x = np.asarray_chkfinite(x)
        return sum(x**2)


    def F2(self,x):  # schw.m
        x = np.asarray_chkfinite(x)
        return sum(abs(x))+prod(abs(x))


    def F3(self,x):
        x = np.asarray_chkfinite(x)
        n = len(x)
        sum=0
        for i in range(n):
            insum = 0
            for j in range(i+1):
                insum += x[j]
            insum *= insum
            sum += insum
        return sum




    def F4(self,x):  # rosen.m
        """ http://en.wikipedia.org/wiki/Rosenbrock_function """
        # a sum of squares, so LevMar (scipy.optimize.leastsq) is pretty good
        x = np.asarray_chkfinite(x)
        x0 = x[:-1]
        x1 = x[1:]
        return sum((1 - x0) **2)+100 * sum((x1 - x0**2)**2)


    def F5(self,x):  # schw.m
        x = np.asarray_chkfinite(x)
        return - sum(x * sin(sqrt(abs(x))))


    def F6(self,x):  # rast.m
        x = np.asarray_chkfinite(x)
        n = len(x)
        return 10*n + sum(x**2 - 10 * cos(2 * pi * x))


    def F7(self,x, a=20, b=0.2, c=2*pi):
        x = np.asarray_chkfinite(x)  # ValueError if any NaN or Inf
        n = len(x)
        s1 = sum(x**2)
        s2 = sum(cos(c * x))
        return -a*exp(-b*sqrt(s1 / n)) - exp(s2 / n) + a + exp(1)


    def F8(self,x, fr=4000):
        x = np.asarray_chkfinite(x)
        n = len(x)
        j = np.arange(1., n+1 )
        s = sum(x**2)
        p = prod(cos(x / sqrt(j)))
        return s/fr - p + 1

    def F9(self,x):
        x = np.asarray_chkfinite(x)
        sum1=0
        sum2=0
        for i in range(len(x)):
            y=self.yi(x[i])
            if(i<len(x)-1):
                sum1=((y-1)**2)*(1+10*(sin(math.pi*self.yi(x[i+1]))**2))
            else:
                sum1+= (y-1)**2
            sum2+= self.u(x[i],10,100,4)
        return (math.pi/len(x))*(10*sin(math.pi*self.yi(x[1]))+sum1)+sum2


    def yi(self,x):
        return 1 + (x+1)/4


    def u(self,x,a,k,m):
        if(x<a):
            return k*(x-a)**m
        elif(x<-a):
            return k*(-x-a)**m
        else:
            return 0


    def F10(self,x):
        a=[[-32,-32],[-16,-32],[0,-32],[16,-32],[32,-32],[-32,-16],[-16,-16],[0,-16],[16,-16],[32,-16],[-32,0],[-16,0],[0,0],[16,0],[32,0],[-32,16],[-16,16],[0,16],[16,16],[32,16],[-32,32],[-16,32],[0,32],[16,32],[32,32]]
        sum=0
        for i in range(25):
            sumint=0
            for j in range(2):
                sumint+=(x[j]-a[i][j])**6
            sumint+= i
            sum+=(1/sumint)
        sum+=(1/500)
        sum=(1/sum)
        return sum

    def F11(self,x):
        return 4*x[0]**2-2.1*x[0]**4+(1/3)*x[0]**6+x[0]*x[1]+4*x[1]**4-4*x[1]**2

    def F12(self,x):
        return (x[1]-(5.1/(4*(math.pi**2)))*(x[0]**2)+(5/math.pi)*x[0]-6)**2+10*(1-(1/(8*math.pi)))*math.cos(x[0])+10

    def F13(self,x):
        return (1+((x[0]+x[1]+1)**2)*(19-14*x[0]+3*(x[0]**2)-14*x[1]+6*x[0]*x[1]+3*(x[1]**2)))*(30+((2*x[0]-3*x[1])**2)*(18-32*x[0]+12*(x[0]**2)+48*x[1]-36*x[0]*x[1]+27*(x[1]**2)))

    def F14(self,x):
        a=[[3,10,30],
           [0.1,10,35],
           [3,10,30],
           [0.1,10,30]]
        c=[1,1.2,3,3.2]
        p=[[0.3689,0.1170,0.2673],
           [0.4699,0.4387,0.7470],
           [0.1091,0.8732,0.5547],
           [0.03815,0.5743,0.8828]]
        sum=0
        for i in range(4):
            sumint=0
            for j in range(3):
                sumint+=a[i][j]*((x[j]-p[i][j])**2)
            sumint= -sumint
            sum+= c[i]*math.exp(sumint)
        return -sum

    def F15(self,x):
        a = [[10,3,17,3.5,1.7,8],
             [0.05,10,17,0.1,8,14],
             [3,3.5,1.7,10,17,8],
             [17,8,0.05,10,0.1,14]]
        c = [1, 1.2, 3, 3.2]
        p = [[0.131,0.169,0.556,0.012,0.828,0.588],
             [0.232,0.412,0.830,0.373,0.100,0.999],
             [0.234,0.141,0.352,0.288,0.304,0.665],
             [0.404,0.882,0.873,0.574,0.109,0.038]]
        sum = 0
        for i in range(4):
            sumint = 0
            for j in range(6):
                sumint += a[i][j] * ((x[j] - p[i][j]) ** 2)
            sumint = -sumint
            sum += c[i] * math.exp(sumint)
        return -sum


class Metaheuristic:

    numberIndividual = 0
    numberIteration = 0
    numberField = 1
    consultationFactor = 0.0
    numberOfVariables = 0
    numberOfVariablesToChange = 0
    iteracionMejorFitness = 0
    k1=0.0
    k2=0.0
    poblation = []
    fields = []
    numAcceptedMoves = 0
    numRejectedMoves = 0
    rpd = 0.0

    def __init__(self, numberindividual, numberiteration, iteration, type, worksheet, times, inten, diver):
        self.numberIndividual=numberindividual
        self.numberIteration=numberiteration
        self.iter=iteration
        self.type=type
        self.fitness=Functions(self.type)
        self.ws=worksheet
        self.times=times
        self.inten = inten
        self.diver = diver

    def run(self):
        if self.type=="12":
            self.ws.cell(row=1, column=1, value="Resultados metaheuristica con formula " + self.type)

            self.generateInitialPoblation()
            self.ordenarhienas()
            h = 5
            e = self.generaRandom()
            e[0] = 2 * e[0] * h - h
            e[1] = 2 * e[1] * h - h
            b = self.generaRandom()
            b[0] = 2 * b[0]
            b[1] = 2 * b[1]

            iteration = 0
            ch = []
            mejorRana = self.poblation[0]
            mejorFitness = self.poblation[0]
            self.ws.cell(row=3, column=1, value="Inicial")
            for i in range(1):
                self.ws.cell(row=3, column=i + 2, value=self.opt(self.poblation[i]))

            while iteration < self.numberIteration:

                for i in range(len(self.poblation)):
                    for _ in range(self.inten):
                        nuevaRana=[0,0]
                        pk = self.poblation[i]
                        nuevaRana[0] = mejorRana[0] - (e[0] * ((b[0] * mejorRana[0]) - pk[0]))
                        nuevaRana[1] = mejorRana[1] - (e[1] * ((b[1] * mejorRana[1]) - pk[1]))
                        pk = nuevaRana
                    if self.opt(pk) < self.opt(self.poblation[i]):
                        ch.append(pk)
                    else:
                        pk = self.poblation[i]
                        ch.append(pk)

                for i in range(len(ch)):
                    for _ in range(self.diver):
                        atake = [0,0]
                        actualrana = ch[i]
                        atake[0] = actualrana[0] / len(ch)
                        atake[1] = actualrana[1] / len(ch)
                        pk = atake

                    if self.opt(pk) < self.opt(ch[i]):
                        self.poblation[i] = pk
                    else:
                        self.poblation[i] = ch[i]
                h = 5 - iteration / self.numberIteration
                e = self.generaRandom()
                e[0] = 2 * e[0] * h - h
                e[1] = 2 * e[1] * h - h
                b = self.generaRandom()
                b[0] = 2 * b[0]
                b[1] = 2 * b[1]
                self.ordenarhienas()
                if self.opt(self.poblation[0]) < self.opt(mejorFitness):
                    mejorRana = self.poblation[0]
                    mejorFitness = self.poblation[0]
                    self.iteracionMejorFitness = iteration
                iteration += 1
                ch = []

                #print("iteracion " + str(iteration + (5000 * (self.times - 1))) + " funcion F" + str(self.iter + 1))
                self.ws.cell(row=3 + iteration, column=1, value=iteration)
                for i in range(1):
                    self.ws.cell(row=3 + iteration, column=i + 2, value=self.opt(self.poblation[i]))
            self.ws.cell(row=6, column=3, value="Fitness")
            self.ws.cell(row=5, column=3, value="Valor")
            self.ws.cell(row=7, column=3, value="opt")
            for i in range(len(self.poblation)):
                self.ws.cell(row=6, column=i + 4, value=self.fitness.calculaFitness(self.poblation[i]))
                self.ws.cell(row=5, column=i + 4, value=str(self.poblation[i][0])+"-"+str(self.poblation[i][1]))
                self.ws.cell(row=7, column=i + 4, value=self.opt(self.poblation[i]))
        else:
            self.ws.cell(row=1, column=1, value="Resultados metaheuristica con formula "+self.type)


            self.generateInitialPoblation()
            self.ordenarhienas()
            h=5
            e=self.generaRandom()
            e=2*e*h-h
            b=self.generaRandom()
            b=2*b

            iteration=0
            ch=[]
            mejorRana=self.poblation[0]
            mejorFitness=self.poblation[0]
            self.ws.cell(row=3,column=1,value="Inicial")
            for i in range(1):
                self.ws.cell(row=3,column=i+2,value=self.opt(self.poblation[i]))


            while iteration<self.numberIteration:

                for i in range(len(self.poblation)):
                    pk = self.poblation[i]
                    for _ in range(self.inten):
                        nuevaRana=mejorRana-(e*((b*mejorRana)-pk))
                        pk=nuevaRana
                    if self.opt(pk)<self.opt(self.poblation[i]):
                        ch.append(pk)
                    else:
                        pk=self.poblation[i]
                        ch.append(pk)

                for i in range(len(ch)):
                    pk= self.poblation[i]
                    for _ in range(self.diver):
                        atake=0
                        actualrana=ch[i]
                        atake=actualrana / len(ch)
                        pk=atake

                    if  self.opt(pk)< self.opt(ch[i]) :
                        self.poblation[i]=pk
                    else:
                        self.poblation[i]=ch[i]
                h=5-iteration/self.numberIteration
                e = self.generaRandom()
                e = 2 * e * h - h
                b = self.generaRandom()
                b = 2 * b
                self.ordenarhienas()
                if self.opt(self.poblation[0])<self.opt(mejorFitness):
                    mejorRana=self.poblation[0]
                    mejorFitness=self.poblation[0]
                    self.iteracionMejorFitness=iteration
                iteration += 1
                ch=[]

                #print("iteracion "+str(iteration+(5000*(self.times-1)))+ " funcion F"+str(self.iter+1))
                self.ws.cell(row=3+iteration, column=1, value=iteration)
                for i in range(1):
                    self.ws.cell(row=3+iteration, column=i + 2, value=self.opt(self.poblation[i]))
            self.ws.cell(row=6, column=3, value="Fitness")
            self.ws.cell(row=5, column=3, value="Valor")
            self.ws.cell(row=7, column=3, value="opt")
            for i in range(len(self.poblation)):
                self.ws.cell(row=6, column=i + 4, value=self.fitness.calculaFitness(self.poblation[i]))
                self.ws.cell(row=5, column=i + 4, value=self.poblation[i])
                self.ws.cell(row=7, column=i + 4, value=self.opt(self.poblation[i]))



    def opt(self,x):
        fit=self.fitness.calculaFitness(x)
        if (self.type == "1"):
            return math.fabs(fit-0)
        elif (self.type == "2"):
            return math.fabs(fit-0)
        elif (self.type == "3"):
            return math.fabs(fit-0)
        elif (self.type == "4"):
            return math.fabs(fit-0)
        elif (self.type == "5"):
            return math.fabs(fit+12.569487)
        elif (self.type == "6"):
            return math.fabs(fit-0)
        elif (self.type == "7"):
            return math.fabs(fit-0)
        elif (self.type == "8"):
            return math.fabs(fit-0)
        elif (self.type == "9"):
            return math.fabs(fit-0)
        elif (self.type == "10"):
            return math.fabs(fit-1)
        elif (self.type == "11"):
            return math.fabs(fit+1.0316285)
        elif (self.type == "12"):
            return math.fabs(fit-0.397887)
        elif (self.type == "13"):
            return math.fabs(fit-3)
        elif self.type == "14":
            return math.fabs(fit+3.86)
        else:
            return math.fabs(fit+3.32)



    def ordenarhienas(self):
        orden=False
        i=0
        while (i< len(self.poblation)) and (orden== False) :
            i+=1
            orden=True
            for j in range(len(self.poblation)-1):
                if self.opt(self.poblation[j+1]) < self.opt(self.poblation[j]):
                    orden = False
                    aux = self.poblation[j]
                    self.poblation[j] = self.poblation[j+1]
                    self.poblation[j+1]= aux
        return


    def generaRandom(self):

        if (self.type == "1"):
            randomSolution = random.random() * 200 - 100
        elif (self.type == "2"):
            randomSolution = random.random() * 20 - 10
        elif (self.type == "3"):
            randomSolution = random.random() * 200 - 100
        elif (self.type == "4"):
            randomSolution = random.random() * 60 - 30
        elif (self.type == "5"):
            randomSolution = random.random() * 1000 - 500
        elif (self.type == "6"):
            randomSolution = random.random() * 10.24 - 5.12
        elif (self.type == "7"):
            randomSolution = random.random() * 64 - 32
        elif (self.type == "8"):
            randomSolution = random.random() * 1200 - 600
        elif (self.type == "9"):
            randomSolution = random.random() * 200 - 100
        elif (self.type == "10"):
            randomSolution = random.random() * 65.536 * 2 - 65.536
        elif (self.type == "11"):
            randomSolution = random.random() * 10 - 5
        elif (self.type == "12"):
            randomSolution = [random.random() * 15 - 5, random.random() * 15]
        elif (self.type == "13"):
            randomSolution = random.random() * 4 - 2
        else:
            randomSolution = random.random()
        return randomSolution

    def discretiza(self, x):
        t = random.random()
        if t < x:
            return 1
        else:
            return 0


    def transforma(self,var):
        t=math.tan(var)
        if t<0:
            t *= -1
        return t



    def generateInitialPoblation(self):
        self.poblation=[]
        for i in range(self.numberIndividual):
            if(self.type=="1"):
                randomSolution=random.random()*200 - 100
            elif(self.type=="2"):
                randomSolution = random.random() * 20 - 10
            elif (self.type == "3"):
                randomSolution = random.random() * 200 - 100
            elif (self.type == "4"):
                randomSolution = random.random() * 60 - 30
            elif (self.type == "5"):
                randomSolution = random.random() * 1000 - 500
            elif (self.type == "6"):
                randomSolution = random.random() * 10.24 - 5.12
            elif (self.type == "7"):
                randomSolution = random.random() * 64 - 32
            elif (self.type == "8"):
                randomSolution = random.random() * 1200 - 600
            elif (self.type == "9"):
                randomSolution = random.random() * 200 - 100
            elif (self.type == "10"):
                randomSolution = random.random() * 65.536*2 - 65.536
            elif (self.type == "11"):
                randomSolution = random.random() * 10 - 5
            elif (self.type == "12"):
                randomSolution = [random.random() * 15 - 5 , random.random()*15]
            elif (self.type == "13"):
                randomSolution = random.random() * 4 - 2
            else:
                randomSolution = random.random()
            self.poblation.append(randomSolution)



numberPoblation = 25
numberIterations = 5000

# parser = argparse.ArgumentParser()
# parser.add_argument("-f", "--function", help="function")
# parser.add_argument("--seed", help="semilla")
# parser.add_argument("-int", "--intensification", type=int)
# parser.add_argument("-div", "--diversification", type=int)

# args = parser.parse_args()

# random.seed(args.seed)
# metaheuristic= Metaheuristic(numberPoblation, numberIterations, 0, args.function, None, 0, args.intensification, args.diversification)

# metaheuristic.run()
numberIterationsForStatistics=15
iterationsForStatistics=0
intensification = 8
diver = 6

while iterationsForStatistics<numberIterationsForStatistics:
    wb=op.Workbook()
    ws1=wb.active
    ws1.title="iteracion 1"
    ws2=wb.create_sheet("Iteracion 2")
    ws3 = wb.create_sheet("Iteracion 3")
    ws4 = wb.create_sheet("Iteracion 4")
    ws5 = wb.create_sheet("Iteracion 5")
    ws6 = wb.create_sheet("Iteracion 6")
    ws7 = wb.create_sheet("Iteracion 7")
    ws8 = wb.create_sheet("Iteracion 8")
    ws9 = wb.create_sheet("Iteracion 9")
    ws10 = wb.create_sheet("Iteracion 10")
    metaheuristic= Metaheuristic(numberPoblation, numberIterations,iterationsForStatistics, str(iterationsForStatistics+1),ws1,1, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws2,2, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws3,3, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws4,4, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws5,5, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws6,6, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws7,7, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws8,8, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws9,9, intensification, diver)
    metaheuristic.run()
    metaheuristic = Metaheuristic(numberPoblation, numberIterations, iterationsForStatistics, str(iterationsForStatistics+1), ws10,10, intensification, diver)
    metaheuristic.run()
    iterationsForStatistics += 1
    wb.save("ResultsF"+str(iterationsForStatistics)+".xlsx")
